"""產生標本批次匯入用的空白範本 lanyang_specimen_import_template.xlsx。

用法：
    python manage.py make_import_template
    python manage.py make_import_template --output 路徑/檔名.xlsx

Sheet1「標本資料」：標題列＋2 列範例，供館員填寫。
Sheet2「填寫說明」：列出保育等級／生命階段的可填中文值與規則。
標題列與對照值直接沿用 collection.resources 的定義，避免範本與匯入邏輯脫節。
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from django.core.management.base import BaseCommand

from collection.resources import (
    CONSERVATION_MAP, LIFE_STAGE_MAP, TEMPLATE_HEADERS,
)

DEFAULT_FILENAME = "lanyang_specimen_import_template.xlsx"

# 兩列示範資料：一列有學名、一列留空（示範系統自動標「[待查證]」）
EXAMPLE_ROWS = [
    ["小白鷺", "Egretta garzetta", "成鳥", "一般類",
     "2026-05-01", "宜蘭縣蘇澳鎮", "王小明", "翼受傷後不治", "IMG_0001.jpg"],
    ["不知名的小型鷸", "", "幼鳥", "待查證",
     "2026-05-03", "宜蘭縣壯圍鄉", "陳小華", "待鑑定", "IMG_0002.jpg"],
]


class Command(BaseCommand):
    help = "產生標本批次匯入用的空白範本 xlsx（含填寫說明工作表）。"

    def add_arguments(self, parser):
        parser.add_argument(
            "--output", default=DEFAULT_FILENAME,
            help=f"輸出檔路徑（預設：{DEFAULT_FILENAME}）",
        )

    def handle(self, *args, **options):
        output = options["output"]
        workbook = Workbook()

        self._build_data_sheet(workbook.active)
        self._build_help_sheet(workbook.create_sheet("填寫說明"))

        workbook.save(output)
        self.stdout.write(self.style.SUCCESS(f"已產生範本：{output}"))

    def _build_data_sheet(self, sheet):
        sheet.title = "標本資料"
        header_font = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="DBEAFE")  # 淺藍

        sheet.append(TEMPLATE_HEADERS)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in EXAMPLE_ROWS:
            sheet.append(row)

        # 欄寬便於閱讀
        widths = [16, 24, 12, 14, 14, 20, 12, 24, 16]
        for idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[sheet.cell(row=1, column=idx).column_letter].width = width
        sheet.freeze_panes = "A2"

    def _build_help_sheet(self, sheet):
        title_font = Font(bold=True, size=12)
        bold = Font(bold=True)

        def add(text="", font=None):
            sheet.append([text])
            if font:
                sheet.cell(row=sheet.max_row, column=1).font = font

        add("蘭陽博物館 標本批次匯入 — 填寫說明", title_font)
        add()
        add("【填寫規則】", bold)
        add("1. 每一列代表一件標本。")
        add("2. 「中文名」為必填。")
        add("3. 「學名」可留空；留空時系統會自動建立物種並標示「[待查證] 中文名」，日後再補正。")
        add("4. 「採集日期」請用 YYYY-MM-DD 格式（例：2026-05-01），可留空。")
        add("5. 採集地點／採集者可留空；三項採集資訊全部留空時，該標本不會建立採集事件。")
        add("6. 「照片檔名」「標本狀況」等沒有專屬欄位的資訊會彙整寫入標本備註。")
        add("7. 同一次匯入中，多列填相同「中文名」只會建立一筆物種。")
        add()
        add("【保育等級 可填的中文值】", bold)
        for label in CONSERVATION_MAP:
            add(f"　{label}")
        add("　（留空＝待查證）")
        add()
        add("【生命階段 可填的中文值】", bold)
        for label in LIFE_STAGE_MAP:
            add(f"　{label}")
        add("　（留空＝不填）")

        sheet.column_dimensions["A"].width = 60
